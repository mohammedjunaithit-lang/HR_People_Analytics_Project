import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import random
from datetime import datetime, timedelta, date

np.random.seed(42)
random.seed(42)

departments = ['Engineering','Sales','HR','Finance','Marketing','Operations','Legal','Product']
locations   = ['Mumbai','Delhi','Bangalore','Chennai','Hyderabad','Pune']
job_levels  = ['L1','L2','L3','L4','L5']
job_titles  = {
    'Engineering': ['Junior Dev','Software Engineer','Senior Engineer','Tech Lead','Engineering Manager'],
    'Sales':       ['Sales Rep','Account Exec','Senior AE','Sales Manager','VP Sales'],
    'HR':          ['HR Coordinator','HR Generalist','HR Manager','HRBP','HR Director'],
    'Finance':     ['Analyst','Senior Analyst','Finance Manager','Controller','CFO'],
    'Marketing':   ['Marketing Coord','Marketing Specialist','Sr. Specialist','Marketing Manager','CMO'],
    'Operations':  ['Ops Analyst','Ops Specialist','Ops Manager','Sr. Ops Manager','VP Ops'],
    'Legal':       ['Legal Analyst','Legal Counsel','Senior Counsel','Legal Manager','General Counsel'],
    'Product':     ['Associate PM','Product Manager','Senior PM','Group PM','VP Product'],
}
genders      = ['Male','Female','Non-Binary']
education    = ['High School','Bachelor','Master','PhD','MBA']
first_names  = ['Aarav','Priya','Rohan','Sneha','Arjun','Ananya','Vikram','Kavya','Rahul','Pooja',
                'Amit','Neha','Suresh','Divya','Karan','Meera','Raj','Swati','Anil','Sunita',
                'Deepak','Rekha','Sanjay','Lakshmi','Nikhil','Asha','Ravi','Geeta','Manish','Radha']
last_names   = ['Sharma','Patel','Singh','Kumar','Gupta','Verma','Joshi','Mehta','Nair','Reddy',
                'Rao','Iyer','Pillai','Bose','Das','Chatterjee','Mishra','Tiwari','Yadav','Saxena']

N = 2000

emp_ids   = [f'EMP{str(i).zfill(4)}' for i in range(1, N+1)]
depts     = np.random.choice(departments, N, p=[0.25,0.20,0.08,0.10,0.12,0.12,0.05,0.08])
levels    = np.random.choice(job_levels,  N, p=[0.30,0.28,0.22,0.14,0.06])
titles    = [job_titles[d][job_levels.index(l)] for d,l in zip(depts, levels)]
locs      = np.random.choice(locations, N, p=[0.20,0.18,0.25,0.12,0.15,0.10])
gend      = np.random.choice(genders, N, p=[0.52,0.44,0.04])
edu       = np.random.choice(education, N, p=[0.05,0.40,0.30,0.10,0.15])
fnames    = [random.choice(first_names) for _ in range(N)]
lnames    = [random.choice(last_names)  for _ in range(N)]
names     = [f'{f} {l}' for f,l in zip(fnames,lnames)]

hire_dates = [date(2015,1,1) + timedelta(days=random.randint(0,3000)) for _ in range(N)]

base_salary = {
    'Engineering': {'L1':400000,'L2':600000,'L3':900000,'L4':1400000,'L5':2000000},
    'Sales':       {'L1':350000,'L2':500000,'L3':750000,'L4':1100000,'L5':1800000},
    'HR':          {'L1':300000,'L2':450000,'L3':650000,'L4':950000, 'L5':1400000},
    'Finance':     {'L1':380000,'L2':560000,'L3':800000,'L4':1200000,'L5':1900000},
    'Marketing':   {'L1':320000,'L2':480000,'L3':700000,'L4':1050000,'L5':1600000},
    'Operations':  {'L1':280000,'L2':420000,'L3':620000,'L4':920000, 'L5':1350000},
    'Legal':       {'L1':420000,'L2':640000,'L3':950000,'L4':1500000,'L5':2200000},
    'Product':     {'L1':450000,'L2':680000,'L3':1000000,'L4':1550000,'L5':2300000},
}
salaries = [int(base_salary[d][l] * np.random.uniform(0.92, 1.12)) for d,l in zip(depts,levels)]

perf_ratings = np.random.choice([1,2,3,4,5], N, p=[0.05,0.10,0.40,0.30,0.15])
attrition_prob = [0.25 if r<=2 else (0.12 if r==3 else 0.04) for r in perf_ratings]
attrition_flag = [np.random.random() < p for p in attrition_prob]

exit_dates = []
for i in range(N):
    if attrition_flag[i]:
        hd = hire_dates[i]
        ed = hd + timedelta(days=random.randint(180, 2000))
        if ed > date(2024,12,31): ed = date(2024,12,31)
        exit_dates.append(ed)
    else:
        exit_dates.append(None)

emp_df = pd.DataFrame({
    'Employee_ID':   emp_ids,
    'Full_Name':     names,
    'Gender':        gend,
    'Department':    depts,
    'Job_Title':     titles,
    'Job_Level':     levels,
    'Location':      locs,
    'Education':     edu,
    'Hire_Date':     hire_dates,
    'Exit_Date':     exit_dates,
    'Salary_INR':    salaries,
    'Performance_Rating': perf_ratings,
    'Attrition':     ['Yes' if a else 'No' for a in attrition_flag],
    'Age':           np.random.randint(22, 58, N),
    'Years_Experience': [random.randint(1, 25) for _ in range(N)],
    'Manager_ID':    [f'EMP{str(random.randint(1,200)).zfill(4)}' for _ in range(N)],
    'Training_Hours': np.random.randint(8, 80, N),
    'Satisfaction_Score': np.round(np.random.uniform(2.0, 5.0, N), 1),
    'Overtime_Flag': np.random.choice(['Yes','No'], N, p=[0.35,0.65]),
    'Remote_Flag':   np.random.choice(['Yes','No'], N, p=[0.40,0.60]),
})

# Payroll table
months = pd.date_range('2023-01-01', '2024-12-01', freq='MS')
payroll_rows = []
for _, emp in emp_df.sample(500).iterrows():
    for m in months:
        hra       = int(emp['Salary_INR'] * 0.40 / 12)
        basic     = int(emp['Salary_INR'] * 0.50 / 12)
        allowance = int(emp['Salary_INR'] * 0.10 / 12)
        pf        = int(basic * 0.12)
        tax       = int(emp['Salary_INR'] * 0.10 / 12)
        net       = basic + hra + allowance - pf - tax
        payroll_rows.append({
            'Employee_ID':   emp['Employee_ID'],
            'Month':         m.strftime('%Y-%m'),
            'Basic_INR':     basic,
            'HRA_INR':       hra,
            'Allowance_INR': allowance,
            'PF_Deduction':  pf,
            'Tax_Deduction': tax,
            'Net_Pay_INR':   net,
            'Department':    emp['Department'],
        })
payroll_df = pd.DataFrame(payroll_rows)

# Attendance table
attendance_rows = []
for _, emp in emp_df.sample(500).iterrows():
    for month_offset in range(24):
        m_date = date(2023,1,1) + timedelta(days=30*month_offset)
        present = random.randint(18, 23)
        leaves  = random.randint(0, 4)
        absent  = 23 - present - leaves
        attendance_rows.append({
            'Employee_ID': emp['Employee_ID'],
            'Month':       m_date.strftime('%Y-%m'),
            'Present_Days': present,
            'Leave_Days':   leaves,
            'Absent_Days':  max(0, absent),
            'Late_Count':   random.randint(0, 5),
            'Overtime_Hrs': random.randint(0, 40),
        })
attendance_df = pd.DataFrame(attendance_rows)

# Recruitment table
recruitment_rows = []
for dept in departments:
    for q in range(1,9):
        year = 2023 + (q-1)//4
        qnum = ((q-1) % 4) + 1
        applied   = random.randint(50, 300)
        shortlist = int(applied * random.uniform(0.2, 0.4))
        interviewed = int(shortlist * random.uniform(0.4, 0.7))
        offered   = int(interviewed * random.uniform(0.3, 0.6))
        hired     = int(offered * random.uniform(0.6, 0.9))
        cost      = hired * random.randint(15000, 40000)
        ttf       = random.randint(20, 75)
        recruitment_rows.append({
            'Department':        dept,
            'Year':              year,
            'Quarter':           f'Q{qnum}',
            'Applications':      applied,
            'Shortlisted':       shortlist,
            'Interviewed':       interviewed,
            'Offers_Made':       offered,
            'Hired':             hired,
            'Cost_Per_Hire_INR': cost // max(hired,1),
            'Time_To_Fill_Days': ttf,
        })
recruitment_df = pd.DataFrame(recruitment_rows)

print("Data generated OK")
print(f"  Employees:   {len(emp_df)}")
print(f"  Payroll:     {len(payroll_df)}")
print(f"  Attendance:  {len(attendance_df)}")
print(f"  Recruitment: {len(recruitment_df)}")

# ─── Build the Excel workbook ────────────────────────────────────────────────
wb = Workbook()

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")
NORM_FONT  = Font(name="Arial", size=10)
BOLD_FONT  = Font(name="Arial", bold=True, size=10)
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center")
THIN       = Side(style="thin", color="BFBFBF")
THIN_BORDER= Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ACCENT_FILLS = {
    'teal':   PatternFill("solid", fgColor="1D9E75"),
    'purple': PatternFill("solid", fgColor="7F77DD"),
    'amber':  PatternFill("solid", fgColor="EF9F27"),
    'coral':  PatternFill("solid", fgColor="D85A30"),
    'blue':   PatternFill("solid", fgColor="378ADD"),
}

def write_sheet(ws, df, sheet_title, col_widths=None):
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False
    # Title row
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
    title_cell = ws['A1']
    title_cell.value    = sheet_title
    title_cell.font     = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill     = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment= CENTER
    # Header row
    ws.row_dimensions[2].height = 22
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=col.replace('_',' '))
        c.font      = HDR_FONT
        c.fill      = PatternFill("solid", fgColor="2E75B6")
        c.alignment = CENTER
        c.border    = THIN_BORDER
    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), 3):
        fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        ws.row_dimensions[ri].height = 16
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = NORM_FONT
            c.fill      = fill
            c.alignment = LEFT
            c.border    = THIN_BORDER
    # Column widths
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    else:
        for ci, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max())
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 30)
    # Freeze panes
    ws.freeze_panes = 'A3'
    return ws

# ── Sheet 1: Employee Master ─────────────────────────────────────────────────
ws1 = wb.active
write_sheet(ws1, emp_df, "Employee Master")

# ── Sheet 2: Payroll ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet()
write_sheet(ws2, payroll_df, "Payroll Data")

# ── Sheet 3: Attendance ──────────────────────────────────────────────────────
ws3 = wb.create_sheet()
write_sheet(ws3, attendance_df, "Attendance Data")

# ── Sheet 4: Recruitment ─────────────────────────────────────────────────────
ws4 = wb.create_sheet()
write_sheet(ws4, recruitment_df, "Recruitment Data")

# ── Sheet 5: KPI Summary (with formulas + charts) ───────────────────────────
ws5 = wb.create_sheet("KPI Summary")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 20

ws5.row_dimensions[1].height = 36
ws5.merge_cells('A1:D1')
t = ws5['A1']
t.value     = "HR People Analytics — KPI Summary"
t.font      = Font(name="Arial", bold=True, size=16, color="FFFFFF")
t.fill      = PatternFill("solid", fgColor="1F4E79")
t.alignment = CENTER

kpi_headers = ['Metric', 'Value', 'Target', 'Status']
for ci, h in enumerate(kpi_headers, 1):
    c = ws5.cell(row=2, column=ci, value=h)
    c.font      = HDR_FONT
    c.fill      = PatternFill("solid", fgColor="2E75B6")
    c.alignment = CENTER
    c.border    = THIN_BORDER

total_emp   = len(emp_df)
attrition_n = emp_df['Attrition'].value_counts().get('Yes', 0)
attrition_r = round(attrition_n / total_emp * 100, 1)
avg_salary  = int(emp_df['Salary_INR'].mean())
avg_perf    = round(emp_df['Performance_Rating'].mean(), 2)
avg_tenure  = round((emp_df['Hire_Date'].apply(lambda x: (date(2024,12,31)-x).days/365)).mean(), 1)
avg_sat     = round(emp_df['Satisfaction_Score'].mean(), 2)
avg_train   = round(emp_df['Training_Hours'].mean(), 1)

kpis = [
    ("Total Employees",       total_emp,      2000,  "On Track"),
    ("Attrition Rate (%)",    attrition_r,    10.0,  "Needs Attention" if attrition_r>10 else "On Track"),
    ("Avg Salary (INR)",      avg_salary,     700000,"On Track"),
    ("Avg Performance",       avg_perf,       3.5,   "On Track" if avg_perf>=3.5 else "Below Target"),
    ("Avg Tenure (Yrs)",      avg_tenure,     3.0,   "On Track"),
    ("Avg Satisfaction",      avg_sat,        3.5,   "On Track" if avg_sat>=3.5 else "Below Target"),
    ("Avg Training Hrs",      avg_train,      40.0,  "On Track" if avg_train>=40 else "Below Target"),
    ("Female Ratio (%)",      round(len(emp_df[emp_df['Gender']=='Female'])/total_emp*100,1), 40.0, "On Track"),
]

status_colors = {"On Track": "1D9E75", "Needs Attention": "EF9F27", "Below Target": "D85A30"}
for ri, (metric, val, target, status) in enumerate(kpis, 3):
    row_fill = PatternFill("solid", fgColor="EBF3FB") if ri%2==0 else PatternFill("solid", fgColor="FFFFFF")
    for ci in range(1, 5):
        c = ws5.cell(row=ri, column=ci)
        c.fill   = row_fill
        c.border = THIN_BORDER
        c.font   = NORM_FONT
        c.alignment = CENTER
    ws5.cell(row=ri, column=1, value=metric).alignment = LEFT
    ws5.cell(row=ri, column=2, value=val)
    ws5.cell(row=ri, column=3, value=target)
    sc = ws5.cell(row=ri, column=4, value=status)
    sc.font = Font(name="Arial", bold=True, size=10, color=status_colors[status])

# Dept breakdown table
ws5.cell(row=13, column=1, value="Department Breakdown").font = Font(name="Arial", bold=True, size=12, color="1F4E79")
dept_headers = ['Department','Head Count','Avg Salary','Attrition %','Avg Rating']
for ci, h in enumerate(dept_headers, 1):
    c = ws5.cell(row=14, column=ci, value=h)
    c.font = HDR_FONT; c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = CENTER; c.border = THIN_BORDER

dept_stats = emp_df.groupby('Department').agg(
    Head_Count=('Employee_ID','count'),
    Avg_Salary=('Salary_INR','mean'),
    Attrition_pct=('Attrition', lambda x: round((x=='Yes').sum()/len(x)*100,1)),
    Avg_Rating=('Performance_Rating','mean')
).reset_index()

for ri2, row in enumerate(dept_stats.itertuples(index=False), 15):
    fill2 = PatternFill("solid", fgColor="EBF3FB") if ri2%2==0 else PatternFill("solid", fgColor="FFFFFF")
    ws5.cell(row=ri2, column=1, value=row.Department).fill = fill2
    ws5.cell(row=ri2, column=2, value=row.Head_Count).fill = fill2
    ws5.cell(row=ri2, column=3, value=int(row.Avg_Salary)).fill = fill2
    ws5.cell(row=ri2, column=4, value=row.Attrition_pct).fill = fill2
    ws5.cell(row=ri2, column=5, value=round(row.Avg_Rating,2)).fill = fill2
    for ci3 in range(1,6):
        ws5.cell(row=ri2, column=ci3).border = THIN_BORDER
        ws5.cell(row=ri2, column=ci3).font   = NORM_FONT
        ws5.cell(row=ri2, column=ci3).alignment = CENTER

ws5.column_dimensions['E'].width = 14

# Bar chart: Headcount by dept
chart1 = BarChart()
chart1.type   = "col"
chart1.title  = "Headcount by Department"
chart1.y_axis.title = "Employees"
chart1.x_axis.title = "Department"
chart1.style  = 10
chart1.width  = 18; chart1.height = 12
data_ref  = Reference(ws5, min_col=2, min_row=14, max_row=14+len(dept_stats))
cats_ref  = Reference(ws5, min_col=1, min_row=15, max_row=14+len(dept_stats))
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
ws5.add_chart(chart1, "G3")

# Pie chart: Attrition Yes/No
ws5.cell(row=30, column=1, value="Attrition").font = HDR_FONT
ws5.cell(row=31, column=1, value="Yes"); ws5.cell(row=31, column=2, value=attrition_n)
ws5.cell(row=32, column=1, value="No");  ws5.cell(row=32, column=2, value=total_emp-attrition_n)
pie = PieChart()
pie.title  = "Attrition Distribution"
pie.style  = 10
pie.width  = 14; pie.height = 10
pdata = Reference(ws5, min_col=2, min_row=31, max_row=32)
pcats = Reference(ws5, min_col=1, min_row=31, max_row=32)
pie.add_data(pdata)
pie.set_categories(pcats)
slice1 = DataPoint(idx=0); slice1.graphicalProperties.solidFill = "D85A30"
slice2 = DataPoint(idx=1); slice2.graphicalProperties.solidFill = "1D9E75"
pie.series[0].data_points = [slice1, slice2]
ws5.add_chart(pie, "G20")

wb.save("/home/claude/hr_project/data/HR_People_Analytics_Data.xlsx")
print("\nExcel workbook saved!")
